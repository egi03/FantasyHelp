from rest_framework import viewsets, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView
from rest_framework.views import APIView
from django.core.cache import cache
from django.db.models import QuerySet
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_headers
from django.http import HttpResponse
from django.utils import timezone
from typing import Dict, Any, Optional, List, Type
import structlog
import uuid

from .permissions import BasePermission
from .pagination import StandardResultsSetPagination
from .throttling import BaseRateThrottle
from .exceptions import ValidationError, NotFoundError
from .utils import measure_time, get_client_ip

logger = structlog.get_logger(__name__)


class RequestIdMixin:
    """Add unique request ID for tracking and debugging"""

    def dispatch(self, request, *args, **kwargs):
        # Generate unique request ID
        request.id = str(uuid.uuid4())

        # Add to response headers
        response = super().dispatch(request, *args, **kwargs)
        if hasattr(response, '__setitem__'):
            response['X-Request-ID'] = request.id

        return response


class LoggingMixin:
    """Add comprehensive request/response logging"""

    def dispatch(self, request, *args, **kwargs):
        start_time = timezone.now()

        # Log request
        logger.info(
            "API request started",
            method=request.method,
            path=request.path,
            user_id=request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
            ip_address=get_client_ip(request),
            request_id=getattr(request, 'id', None),
        )

        # Process request
        response = super().dispatch(request, *args, **kwargs)

        # Calculate duration
        duration = (timezone.now() - start_time).total_seconds()

        # Log response
        logger.info(
            "API request completed",
            method=request.method,
            path=request.path,
            status_code=response.status_code,
            duration_seconds=duration,
            request_id=getattr(request, 'id', None),
        )

        return response


class CachingMixin:
    """Add intelligent caching capabilities"""

    cache_timeout = 300  # 5 minutes default
    cache_key_prefix = None
    vary_on_user = False

    def get_cache_key(self, request, *args, **kwargs) -> str:
        """Generate cache key for the request"""
        key_parts = [
            self.cache_key_prefix or self.__class__.__name__.lower(),
            request.path,
            request.GET.urlencode(),
        ]

        if self.vary_on_user and hasattr(request, 'user') and request.user.is_authenticated:
            key_parts.append(f"user:{request.user.id}")

        return ":".join(str(part) for part in key_parts if part)

    def get_cached_response(self, request, *args, **kwargs) -> Optional[Response]:
        """Get cached response if available"""
        if not hasattr(self, 'cache_timeout') or self.cache_timeout <= 0:
            return None

        cache_key = self.get_cache_key(request, *args, **kwargs)
        cached_data = cache.get(cache_key)

        if cached_data is not None:
            logger.debug("Cache hit", cache_key=cache_key)
            return Response(cached_data)

        return None

    def cache_response(self, request, response, *args, **kwargs):
        """Cache the response data"""
        if not hasattr(self, 'cache_timeout') or self.cache_timeout <= 0:
            return

        if response.status_code == 200:
            cache_key = self.get_cache_key(request, *args, **kwargs)
            cache.set(cache_key, response.data, self.cache_timeout)
            logger.debug("Response cached", cache_key=cache_key)


class OptimizedQuerysetMixin:
    """Optimize querysets for better performance"""

    select_related_fields = []
    prefetch_related_fields = []

    def get_queryset(self):
        """Apply optimizations to queryset"""
        queryset = super().get_queryset()

        # Apply select_related
        if self.select_related_fields:
            queryset = queryset.select_related(*self.select_related_fields)

        # Apply prefetch_related
        if self.prefetch_related_fields:
            queryset = queryset.prefetch_related(*self.prefetch_related_fields)

        return queryset

    def optimize_queryset_for_action(self, queryset: QuerySet) -> QuerySet:
        """Optimize queryset based on the current action"""
        if hasattr(self, 'action'):
            action_optimizations = getattr(self, f'optimize_for_{self.action}', None)
            if callable(action_optimizations):
                return action_optimizations(queryset)

        return queryset


class BaseAPIView(RequestIdMixin, LoggingMixin, CachingMixin, APIView):
    """
    Base API view with common functionality
    Includes logging, caching, and error handling
    """

    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [BaseRateThrottle]

    def handle_exception(self, exc):
        """Enhanced exception handling with logging"""
        logger.error(
            "API view exception",
            exception=str(exc),
            exception_type=exc.__class__.__name__,
            view=self.__class__.__name__,
            request_id=getattr(self.request, 'id', None),
        )

        return super().handle_exception(exc)

    @measure_time
    def dispatch(self, request, *args, **kwargs):
        """Enhanced dispatch with performance monitoring"""
        return super().dispatch(request, *args, **kwargs)


class BaseModelViewSet(RequestIdMixin, LoggingMixin, OptimizedQuerysetMixin,
                      CachingMixin, viewsets.ModelViewSet):
    """
    Enhanced model viewset with optimizations and common functionality
    """

    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [BaseRateThrottle]
    pagination_class = StandardResultsSetPagination

    def get_serializer_context(self):
        """Add extra context to serializers"""
        context = super().get_serializer_context()
        context.update({
            'request_id': getattr(self.request, 'id', None),
            'user_ip': get_client_ip(self.request),
        })
        return context

    def get_queryset(self):
        """Apply optimizations and filters"""
        queryset = super().get_queryset()
        queryset = self.optimize_queryset_for_action(queryset)
        return queryset

    def list(self, request, *args, **kwargs):
        """Optimized list with caching"""
        # Try cache first
        cached_response = self.get_cached_response(request, *args, **kwargs)
        if cached_response:
            return cached_response

        # Generate response
        response = super().list(request, *args, **kwargs)

        # Cache successful responses
        self.cache_response(request, response, *args, **kwargs)

        return response

    def retrieve(self, request, *args, **kwargs):
        """Optimized retrieve with caching"""
        cached_response = self.get_cached_response(request, *args, **kwargs)
        if cached_response:
            return cached_response

        response = super().retrieve(request, *args, **kwargs)
        self.cache_response(request, response, *args, **kwargs)

        return response

    def perform_create(self, serializer):
        """Enhanced create with logging"""
        instance = serializer.save()

        logger.info(
            "Object created",
            model=instance.__class__.__name__,
            pk=str(instance.pk),
            user_id=self.request.user.id if self.request.user.is_authenticated else None,
        )

        # Invalidate related caches
        self.invalidate_related_caches(instance)

    def perform_update(self, serializer):
        """Enhanced update with logging"""
        instance = serializer.save()

        logger.info(
            "Object updated",
            model=instance.__class__.__name__,
            pk=str(instance.pk),
            user_id=self.request.user.id if self.request.user.is_authenticated else None,
        )

        self.invalidate_related_caches(instance)

    def perform_destroy(self, instance):
        """Enhanced delete with logging"""
        model_name = instance.__class__.__name__
        pk = str(instance.pk)

        logger.info(
            "Object deleted",
            model=model_name,
            pk=pk,
            user_id=self.request.user.id if self.request.user.is_authenticated else None,
        )

        self.invalidate_related_caches(instance)
        super().perform_destroy(instance)

    def invalidate_related_caches(self, instance):
        """Invalidate caches related to the instance"""
        # Override in subclasses to implement specific cache invalidation
        pass

    @action(detail=False, methods=['get'])
    def metadata(self, request):
        """Get metadata about the viewset"""
        return Response({
            'model': self.get_queryset().model.__name__,
            'total_count': self.get_queryset().count(),
            'fields': list(self.get_serializer().fields.keys()),
            'permissions': [p.__name__ for p in self.get_permissions()],
            'throttles': [t.__class__.__name__ for t in self.get_throttles()],
        })


class OptimizedListView(LoggingMixin, CachingMixin, OptimizedQuerysetMixin, ListAPIView):
    """
    Optimized list view for read-heavy endpoints
    """

    cache_timeout = 600  # 10 minutes
    pagination_class = StandardResultsSetPagination

    def list(self, request, *args, **kwargs):
        cached_response = self.get_cached_response(request, *args, **kwargs)
        if cached_response:
            return cached_response

        response = super().list(request, *args, **kwargs)
        self.cache_response(request, response, *args, **kwargs)

        return response


class OptimizedRetrieveView(LoggingMixin, CachingMixin, OptimizedQuerysetMixin, RetrieveAPIView):
    """
    Optimized retrieve view for read-heavy endpoints
    """

    cache_timeout = 1800  # 30 minutes

    def retrieve(self, request, *args, **kwargs):
        cached_response = self.get_cached_response(request, *args, **kwargs)
        if cached_response:
            return cached_response

        response = super().retrieve(request, *args, **kwargs)
        self.cache_response(request, response, *args, **kwargs)

        return response


class BulkActionMixin:
    """Add bulk action capabilities to viewsets"""

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple objects in bulk"""
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)

        instances = serializer.save()

        logger.info(
            "Bulk create completed",
            model=self.get_queryset().model.__name__,
            count=len(instances),
            user_id=request.user.id if request.user.is_authenticated else None,
        )

        return Response(
            self.get_serializer(instances, many=True).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['patch'])
    def bulk_update(self, request):
        """Update multiple objects in bulk"""
        if not isinstance(request.data, list):
            raise ValidationError("Expected a list of objects")

        updated_objects = []
        errors = []

        for item_data in request.data:
            if 'id' not in item_data:
                errors.append({'error': 'Missing id field', 'data': item_data})
                continue

            try:
                instance = self.get_queryset().get(pk=item_data['id'])
                serializer = self.get_serializer(instance, data=item_data, partial=True)

                if serializer.is_valid():
                    updated_objects.append(serializer.save())
                else:
                    errors.append({'id': item_data['id'], 'errors': serializer.errors})

            except self.get_queryset().model.DoesNotExist:
                errors.append({'error': f"Object with id {item_data['id']} not found"})

        response_data = {
            'updated': len(updated_objects),
            'errors': len(errors),
        }

        if updated_objects:
            response_data['objects'] = self.get_serializer(updated_objects, many=True).data

        if errors:
            response_data['error_details'] = errors

        logger.info(
            "Bulk update completed",
            model=self.get_queryset().model.__name__,
            updated=len(updated_objects),
            errors=len(errors),
        )

        return Response(response_data)

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        """Delete multiple objects in bulk"""
        ids = request.data.get('ids', [])

        if not ids:
            raise ValidationError("No IDs provided")

        queryset = self.get_queryset().filter(pk__in=ids)
        count = queryset.count()

        queryset.delete()

        logger.info(
            "Bulk delete completed",
            model=self.get_queryset().model.__name__,
            count=count,
            user_id=request.user.id if request.user.is_authenticated else None,
        )

        return Response({'deleted': count})


class ReadOnlyModelViewSet(BaseModelViewSet):
    """
    Model viewset that only allows read operations
    """

    http_method_names = ['get', 'head', 'options']

    def create(self, request, *args, **kwargs):
        return Response(
            {'error': 'Create operation not allowed'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

    def update(self, request, *args, **kwargs):
        return Response(
            {'error': 'Update operation not allowed'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

    def destroy(self, request, *args, **kwargs):
        return Response(
            {'error': 'Delete operation not allowed'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


class SearchMixin:
    """Add advanced search capabilities"""

    search_fields = []

    @action(detail=False, methods=['get'])
    def search(self, request):
        """Advanced search endpoint"""
        query = request.query_params.get('q', '')

        if not query:
            return Response({'error': 'Query parameter "q" is required'})

        if len(query) < 2:
            return Response({'error': 'Query must be at least 2 characters'})

        queryset = self.get_queryset()

        # Apply search filters
        if hasattr(self, 'get_search_queryset'):
            queryset = self.get_search_queryset(queryset, query)
        else:
            # Default search implementation
            from django.db.models import Q
            search_q = Q()

            for field in self.search_fields:
                search_q |= Q(**{f"{field}__icontains": query})

            queryset = queryset.filter(search_q)

        # Paginate results
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class ExportMixin:
    """Add data export capabilities"""

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export data in various formats"""
        format_type = request.query_params.get('format', 'json').lower()

        queryset = self.filter_queryset(self.get_queryset())

        # Limit export size
        max_export_size = getattr(self, 'max_export_size', 10000)
        if queryset.count() > max_export_size:
            return Response(
                {'error': f'Export size exceeds limit of {max_export_size} records'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if format_type == 'csv':
            return self.export_csv(queryset)
        elif format_type == 'excel':
            return self.export_excel(queryset)
        else:
            # Default JSON export
            serializer = self.get_serializer(queryset, many=True)
            return Response({
                'count': queryset.count(),
                'data': serializer.data,
                'exported_at': timezone.now().isoformat(),
            })

    def export_csv(self, queryset):
        """Export data as CSV"""
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.csv"'

        writer = csv.writer(response)

        # Write headers
        serializer = self.get_serializer()
        headers = list(serializer.fields.keys())
        writer.writerow(headers)

        # Write data
        for obj in queryset:
            serializer = self.get_serializer(obj)
            row = [str(serializer.data.get(field, '')) for field in headers]
            writer.writerow(row)

        return response

    def export_excel(self, queryset):
        """Export data as Excel"""
        try:
            import openpyxl
            from django.http import HttpResponse
            from io import BytesIO
        except ImportError:
            return Response(
                {'error': 'Excel export not available - openpyxl not installed'},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.get_export_filename()

        # Write headers
        serializer = self.get_serializer()
        headers = list(serializer.fields.keys())
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Write data
        for row, obj in enumerate(queryset, 2):
            serializer = self.get_serializer(obj)
            for col, field in enumerate(headers, 1):
                worksheet.cell(row=row, column=col, value=serializer.data.get(field, ''))

        # Create response
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.xlsx"'

        return response

    def get_export_filename(self):
        """Get filename for export"""
        model_name = self.get_queryset().model.__name__.lower()
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        return f"{model_name}_export_{timestamp}"


class StatsMixin:
    """Add statistics endpoints"""

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get statistics about the dataset"""
        queryset = self.get_queryset()

        stats = {
            'total_count': queryset.count(),
            'created_today': queryset.filter(
                created_at__date=timezone.now().date()
            ).count() if hasattr(queryset.model, 'created_at') else None,
        }

        # Add custom stats if implemented
        if hasattr(self, 'get_custom_stats'):
            stats.update(self.get_custom_stats(queryset))

        return Response(stats)


class HealthCheckView(APIView):
    """
    Health check endpoint for monitoring
    """

    permission_classes = [permissions.AllowAny]

    def get(self, request):
        """Perform health checks"""
        health_data = {
            'status': 'healthy',
            'timestamp': timezone.now().isoformat(),
            'checks': {}
        }

        # Database check
        try:
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute('SELECT 1')
            health_data['checks']['database'] = 'ok'
        except Exception as e:
            health_data['checks']['database'] = f'error: {str(e)}'
            health_data['status'] = 'unhealthy'

        # Cache check
        try:
            cache.set('health_check', 'ok', 10)
            cache.get('health_check')
            health_data['checks']['cache'] = 'ok'
        except Exception as e:
            health_data['checks']['cache'] = f'error: {str(e)}'
            health_data['status'] = 'degraded' if health_data['status'] == 'healthy' else 'unhealthy'

        # Return appropriate status code
        status_code = status.HTTP_200_OK
        if health_data['status'] == 'unhealthy':
            status_code = status.HTTP_503_SERVICE_UNAVAILABLE

        return Response(health_data, status=status_code)
